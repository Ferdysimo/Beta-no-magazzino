from datetime import datetime, timedelta, timezone
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.database import db
from app.core.security import require_admin, require_admin_or_federico, verify_token
from app.core.time import ROME_TZ
from app.services.analysis import (
    _build_annual_analysis_data,
    _display_media_location,
    _ensure_analysis_integrity,
    _format_italian_long_date,
    _media_code_for_restaurant,
    _prefetch_analysis_order_data,
    _validate_export_year,
    _write_analysis_locale_sheet,
    _write_totali_sheet_for_analysis,
)


router = APIRouter()


@router.get("/admin/media-locali")
async def get_media_locali(token_data: dict = Depends(verify_token)):
    require_admin_or_federico(token_data)

    # Get all restaurants
    restaurants = await db.restaurants.find(
        {"role": "restaurant"},
        {"_id": 0, "password": 0}
    ).to_list(100)

    # Date range: same day last month to today.
    # NB: usare semplicemente `today.replace(month=today.month - 1)` esplode
    # ogni volta che il giorno corrente non esiste nel mese precedente
    # (es. 31 maggio → 31 aprile). Si clampa al massimo numero di giorni
    # del mese di destinazione.
    import calendar
    today = datetime.now(ROME_TZ).replace(hour=23, minute=59, second=59)
    if today.month > 1:
        prev_year, prev_month = today.year, today.month - 1
    else:
        prev_year, prev_month = today.year - 1, 12
    last_day_prev = calendar.monthrange(prev_year, prev_month)[1]
    from_date = today.replace(
        year=prev_year, month=prev_month, day=min(today.day, last_day_prev)
    )

    result = []

    # For each day in range
    current = today.replace(hour=0, minute=0, second=0, microsecond=0)
    from_start = from_date.replace(hour=0, minute=0, second=0, microsecond=0)
    restaurant_ids = [r.get("id") for r in restaurants if r.get("id")]
    source_data = await _prefetch_analysis_order_data(
        restaurant_ids,
        from_start.astimezone(timezone.utc).isoformat(),
        (current + timedelta(days=1)).astimezone(timezone.utc).isoformat(),
    )
    source_counts = source_data["counts"]

    while current >= from_start:
        day_data = {"date": current.strftime("%d/%m/%Y"), "locations": {}}
        date_rome = current.strftime("%Y-%m-%d")

        for rest in restaurants:
            day_data["locations"][rest["location"]] = int(
                source_counts.get((rest["id"], date_rome), 0) or 0
            )

        result.append(day_data)
        current -= timedelta(days=1)

    # Calculate averages per location, EXCLUDING days with 0/empty values
    # so empty cells in the table don't dilute the average.
    averages = {}
    for rest in restaurants:
        loc = rest["location"]
        values = []
        for d in result:
            v = d["locations"].get(loc)
            try:
                v_int = int(v) if v else 0
            except (TypeError, ValueError):
                v_int = 0
            if v_int > 0:
                values.append(v_int)
        averages[loc] = round(sum(values) / len(values), 2) if values else 0

    return {
        "locations": [r["location"] for r in restaurants],
        "averages": averages,
        "days": result
    }


@router.get("/admin/analisi-mensile/export")
async def export_analisi_mensile_excel(year: int = None, token_data: dict = Depends(verify_token)):
    require_admin(token_data)
    selected_year = _validate_export_year(year)
    data = await _build_annual_analysis_data(selected_year)
    _ensure_analysis_integrity(data)

    wb = Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    default_sheet = wb.active
    wb.remove(default_sheet)
    used_titles = set()
    for rest_data in data["restaurants"]:
        _write_analysis_locale_sheet(wb, rest_data, data, used_titles)
    _write_totali_sheet_for_analysis(
        wb,
        data["restaurants"],
        selected_year,
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"analisi_mensile_{selected_year}.xlsx"
    integrity = data.get("integrity") or {}
    warning_counts = integrity.get("warning_counts") or {}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Analysis-Warning-Count": str(len(integrity.get("warnings") or [])),
            "X-Analysis-Missing-Snapshot-Count": str(
                warning_counts.get("pasta_snapshot_missing", 0)
            ),
            "X-Analysis-Manual-Override-Count": str(
                warning_counts.get("manual_override_used", 0)
            ),
            "X-Analysis-Source-Missing-Count": str(
                warning_counts.get("source_orders_missing", 0)
            ),
        },
    )


@router.get("/admin/media-locali/export")
async def export_media_locali_excel(year: int = None, token_data: dict = Depends(verify_token)):
    require_admin_or_federico(token_data)

    selected_year = year or datetime.now(ROME_TZ).year
    if selected_year < 2020 or selected_year > 2100:
        raise HTTPException(status_code=400, detail="Anno non valido")

    restaurants = await db.restaurants.find(
        {"role": "restaurant"},
        {"_id": 0, "password": 0}
    ).to_list(100)

    restaurants = sorted(restaurants, key=lambda r: (r.get("location") or "").lower())
    start = datetime(selected_year, 1, 1, tzinfo=ROME_TZ)
    end = datetime(selected_year, 12, 31, tzinfo=ROME_TZ)
    restaurant_ids = [r.get("id") for r in restaurants if r.get("id")]
    source_data = await _prefetch_analysis_order_data(
        restaurant_ids,
        start.astimezone(timezone.utc).isoformat(),
        (end + timedelta(days=1)).astimezone(timezone.utc).isoformat(),
    )
    source_counts = source_data["counts"]

    rows = []
    current = start
    while current <= end:
        location_values = {}
        date_rome = current.strftime("%Y-%m-%d")

        for rest in restaurants:
            location_values[rest["location"]] = int(
                source_counts.get((rest["id"], date_rome), 0) or 0
            )

        rows.append({
            "date": current,
            "locations": location_values,
        })
        current += timedelta(days=1)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Numeri {selected_year}"

    location_headers = [_display_media_location(r["location"]) for r in restaurants]
    media_headers = [f"MEDIA {_media_code_for_restaurant(r)}" for r in restaurants]
    headers = ["DATA", *location_headers, "TOTALI", *media_headers, "MEDIA T"]
    ws.append(headers)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    red_font = Font(color="D00000", name="Times New Roman", size=12)
    black_font = Font(color="000000", name="Times New Roman", size=12)
    header_font = Font(color="D00000", name="Times New Roman", size=12)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(color="000000", name="Times New Roman", size=12)

    month_values = {r["location"]: [] for r in restaurants}
    today_rome = datetime.now(ROME_TZ)

    for row in rows:
        excel_row = [_format_italian_long_date(row["date"])]
        daily_total = 0
        for rest in restaurants:
            value = int(row["locations"].get(rest["location"]) or 0)
            excel_row.append(value if value > 0 else None)
            daily_total += value
            if value > 0:
                month_values[rest["location"]].append(value)
        excel_row.append(daily_total if daily_total > 0 else None)

        is_month_end = (row["date"] + timedelta(days=1)).month != row["date"].month
        is_completed_month = (
            row["date"].year < today_rome.year
            or (
                row["date"].year == today_rome.year
                and row["date"].month < today_rome.month
            )
        )
        monthly_averages = []
        if is_month_end and is_completed_month:
            for rest in restaurants:
                values = month_values[rest["location"]]
                monthly_averages.append(round(sum(values) / len(values), 1) if values else None)
            valid_monthly_averages = [v for v in monthly_averages if v is not None]
            total_avg = round(sum(valid_monthly_averages), 1) if valid_monthly_averages else None
            excel_row.extend(monthly_averages)
            excel_row.append(total_avg)
            month_values = {r["location"]: [] for r in restaurants}
        elif is_month_end:
            excel_row.extend([None] * (len(restaurants) + 1))
            month_values = {r["location"]: [] for r in restaurants}
        else:
            excel_row.extend([None] * (len(restaurants) + 1))

        ws.append(excel_row)

    max_col = len(headers)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=max_col):
        for idx, cell in enumerate(row, start=1):
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = red_font if 2 <= idx <= (1 + len(restaurants)) else black_font
            if idx >= len(restaurants) + 3 and cell.value is not None:
                cell.number_format = "0.0"

    ws.column_dimensions["A"].width = 34
    for col_idx in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"numeri_locali_{selected_year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
