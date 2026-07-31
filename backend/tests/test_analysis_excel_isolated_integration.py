import os
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

from conftest import run_isolated


if os.environ.get("PASTA_RUN_ISOLATED_INTEGRATION") != "1":
    pytest.skip(
        "Set PASTA_RUN_ISOLATED_INTEGRATION=1 with an isolated DB_NAME",
        allow_module_level=True,
    )

BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.core.database import client, db
from app.services.analysis import (
    _build_annual_analysis_data,
    _ensure_analysis_integrity,
    _write_analysis_locale_sheet,
    _write_totali_sheet_for_analysis,
)
from app.services.report import PASTA_PRICES_MAP


def test_analysis_workbook_counts_only_non_cancelled_orders():
    run_isolated(_exercise_analysis_workbook())


async def _exercise_analysis_workbook():
    db_name = os.environ["DB_NAME"]
    assert db_name.startswith("pastasciutta_refactor_test_")

    flaminio_id = "analysis-flaminio"
    grazie_id = "analysis-grazie"
    day = "2026-01-15"
    manual_day = "2026-01-16"
    cancelled_only_day = "2026-01-17"
    dictionary_snapshot = [
        {"sigla": sigla, "price": price}
        for sigla, price in PASTA_PRICES_MAP.items()
    ]

    await client.drop_database(db_name)
    try:
        await db.restaurants.insert_many([
            {
                "id": flaminio_id,
                "username": "FlaminioTest",
                "location": "Flaminio",
                "role": "restaurant",
            },
            {
                "id": grazie_id,
                "username": "GrazieTest",
                "location": "Grazie",
                "role": "restaurant",
            },
        ])
        await db.pasta_dictionary.insert_many([
            {
                "restaurant_id": flaminio_id,
                "siglas": dictionary_snapshot,
            },
            {
                "restaurant_id": grazie_id,
                "siglas": dictionary_snapshot,
            },
        ])

        await db.orders.insert_one({
            "id": "f-original",
            "restaurant_id": flaminio_id,
            "created_at": "2026-01-15T09:00:00+00:00",
            "order_number": 1,
            "description": "CARB senza pepe",
        })
        await db.archived_orders.insert_many([
            {
                # Same original timestamp and number: this is only an archive copy.
                "id": "f-original-copy",
                "restaurant_id": flaminio_id,
                "created_at": "2026-01-15T09:00:00Z",
                "order_number": 1,
                "description": "CARB senza pepe",
            },
            {
                "id": "f-amat",
                "restaurant_id": flaminio_id,
                "created_at": "2026-01-15T10:00:00+00:00",
                "order_number": 2,
                "description": "AMAT asporto",
            },
            {
                # Number 3 was deliberately reused after the earlier order was deleted.
                "id": "f-ragu-reused",
                "restaurant_id": flaminio_id,
                "created_at": "2026-01-15T12:00:00+00:00",
                "order_number": 3,
                "description": "RAGU",
            },
            {
                # Analytical type intentionally absent from the price snapshot.
                "id": "f-tonno",
                "restaurant_id": flaminio_id,
                "created_at": "2026-01-15T12:30:00+00:00",
                "order_number": 5,
                "description": "TONNO TA",
            },
            {
                "id": "f-manual-source",
                "restaurant_id": flaminio_id,
                "created_at": "2026-01-16T09:00:00+00:00",
                "order_number": 1,
                "description": "CARB",
            },
            {
                "id": "g-cacio",
                "restaurant_id": grazie_id,
                "created_at": "2026-01-15T09:30:00+00:00",
                "order_number": 1,
                "description": "CACIO",
            },
            {
                "id": "g-carzuc",
                "restaurant_id": grazie_id,
                "created_at": "2026-01-15T10:30:00+00:00",
                "order_number": 2,
                "description": "CARZUC",
            },
            {
                "id": "g-amat",
                "restaurant_id": grazie_id,
                "created_at": "2026-01-15T11:30:00+00:00",
                "order_number": 3,
                "description": "AMAT",
            },
        ])
        await db.deletion_logs.insert_many([
            {
                "id": "f-pesto-deleted",
                "restaurant_id": flaminio_id,
                "original_created_at": "2026-01-15T11:00:00+00:00",
                "order_number": 3,
                "description": "PESTO",
            },
            {
                "id": "f-xl-deleted",
                "restaurant_id": flaminio_id,
                "original_created_at": "2026-01-15T13:00:00+00:00",
                "order_number": 4,
                "description": "CARB XL",
            },
            {
                "id": "f-cancelled-only",
                "restaurant_id": flaminio_id,
                "original_created_at": "2026-01-17T09:00:00+00:00",
                "order_number": 1,
                "description": "CARB",
            },
        ])
        await db.archived_deletion_logs.insert_one({
            # Same deleted order in both log collections: count it only once.
            "id": "f-pesto-deleted-copy",
            "restaurant_id": flaminio_id,
            "original_created_at": "2026-01-15T11:00:00Z",
            "order_number": 3,
            "description": "PESTO",
        })
        await db.cash_daily_counts.insert_many([
            {
                "restaurant_id": flaminio_id,
                "date_rome": day,
                # The automatic stored text is stale; source orders must win.
                "paste_text": "1 POM",
                "paste_manual_override": False,
                "pasta_dict_snapshot": dictionary_snapshot,
                "ft": "500+300-20",
                "comments": {"ft": "Tre fatture controllate"},
            },
            {
                "restaurant_id": flaminio_id,
                "date_rome": manual_day,
                # A deliberate manual correction must win, but be reported.
                "paste_text": "1 AMAT",
                "paste_manual_override": True,
                "pasta_dict_snapshot": dictionary_snapshot,
            },
            {
                "restaurant_id": flaminio_id,
                "date_rome": cancelled_only_day,
                # Legacy automatic snapshots used to contain cancelled orders.
                "paste_text": "1 CARB",
                "paste_manual_override": False,
                "pasta_dict_snapshot": dictionary_snapshot,
            },
        ])
        await db.beverage_daily_counts.insert_one({
            "restaurant_id": flaminio_id,
            "date_rome": day,
            "sigla": "AL",
            "mattina": "8+2",
            "inUsc": "1+1",
            "scarti": "1",
            "sera": "4",
            "comments": {
                "inUsc": "Consegna extra",
                "scarti": "Bottiglia rotta",
            },
        })

        data = await _build_annual_analysis_data(2026)
        _ensure_analysis_integrity(data)

        restaurants = {item["location"]: item for item in data["restaurants"]}
        flaminio = restaurants["Flaminio"]
        grazie = restaurants["Grazie"]
        flaminio_day = next(row for row in flaminio["rows"] if row["date_str"] == day)
        flaminio_manual = next(
            row for row in flaminio["rows"] if row["date_str"] == manual_day
        )
        flaminio_cancelled_only = next(
            row for row in flaminio["rows"] if row["date_str"] == cancelled_only_day
        )
        grazie_day = next(row for row in grazie["rows"] if row["date_str"] == day)

        # Independent expected totals: deleted rows stay out and TONNO is analytical.
        assert flaminio_day["source_order_count"] == 4
        assert flaminio_day["deleted_order_count"] == 2
        assert flaminio_day["paste_total_count"] == 4
        assert flaminio_day["paste"]["breakdown"]["CARB"]["count"] == 1
        assert flaminio_day["paste"]["breakdown"]["AMAT"]["count"] == 1
        assert flaminio_day["paste"]["breakdown"]["PESTO"]["count"] == 0
        assert flaminio_day["paste"]["breakdown"]["RAGU"]["count"] == 1
        assert flaminio_day["paste"]["breakdown"]["TONNO"]["count"] == 1
        assert flaminio_day["paste"]["unrecognized_count"] == 0
        assert flaminio_day["cash_raw"]["ft"] == "500+300-20"
        assert flaminio_day["cash_comments"]["ft"] == "Tre fatture controllate"
        assert flaminio_day["beverages"]["AL"]["raw"]["inUsc"] == "1+1"
        assert (
            flaminio_day["beverages"]["AL"]["comments"]["scarti"]
            == "Bottiglia rotta"
        )

        assert grazie_day["source_order_count"] == 3
        assert grazie_day["paste_total_count"] == 3
        assert grazie_day["paste"]["breakdown"]["CACIO"]["count"] == 1
        assert "CARZUC" not in grazie_day["paste"]["breakdown"]
        assert "AMAT" not in grazie_day["paste"]["breakdown"]
        assert grazie_day["paste"]["unrecognized_count"] == 2
        assert grazie_day["paste"]["unrecognized_eur"] == 16

        assert flaminio_manual["source_order_count"] == 1
        assert flaminio_manual["paste"]["breakdown"]["AMAT"]["count"] == 1
        assert flaminio_manual["paste"]["breakdown"]["CARB"]["count"] == 0
        assert flaminio_cancelled_only["source_order_count"] == 0
        assert flaminio_cancelled_only["deleted_order_count"] == 1
        assert flaminio_cancelled_only["paste_total_count"] == 0
        warning_codes = {
            warning["code"] for warning in data["integrity"]["warnings"]
        }
        assert "automatic_snapshot_rebuilt" in warning_codes
        assert "manual_override_used" in warning_codes

        wb = Workbook()
        wb.remove(wb.active)
        used_titles = set()
        for restaurant in data["restaurants"]:
            _write_analysis_locale_sheet(wb, restaurant, data, used_titles)
        _write_totali_sheet_for_analysis(wb, data["restaurants"], 2026)

        flaminio_ws = wb["Flaminio"]
        headers = [cell.value for cell in flaminio_ws[7]]

        def first_column(label):
            return headers.index(label) + 1

        excel_row = 8 + 14  # January 15, with January 1 at row 8.
        assert flaminio_ws.cell(excel_row, first_column("Carb")).value == 1
        assert flaminio_ws.cell(excel_row, first_column("Amat")).value == 1
        assert flaminio_ws.cell(excel_row, first_column("Pesto")).value is None
        assert flaminio_ws.cell(excel_row, first_column("Ragu")).value == 1
        assert flaminio_ws.cell(excel_row, first_column("Tonno")).value == 1
        assert flaminio_ws.cell(excel_row, first_column("Altro")).value is None
        assert flaminio_ws.cell(excel_row, first_column("TOT PIATTI")).value == 4
        ft_cell = flaminio_ws.cell(excel_row, first_column("FT"))
        assert ft_cell.value.startswith("=500+300-20")
        assert 'N("Tre fatture controllate")' in ft_cell.value
        assert any(
            ft_cell.coordinate in validation.cells
            and validation.prompt == "Tre fatture controllate"
            for validation in flaminio_ws.data_validations.dataValidation
        )
        group_headers = [cell.value for cell in flaminio_ws[6]]
        scarichi_cell = flaminio_ws.cell(
            excel_row,
            group_headers.index("SCARICHI") + 1,
        )
        assert scarichi_cell.value.startswith("=1+1")
        assert any(
            scarichi_cell.coordinate in validation.cells
            and validation.prompt == "Consegna extra"
            for validation in flaminio_ws.data_validations.dataValidation
        )
        cancelled_excel_row = 8 + 16
        assert flaminio_ws.cell(
            cancelled_excel_row, first_column("TOT PIATTI")
        ).value is None

        grazie_ws = wb["Grazie"]
        grazie_headers = [cell.value for cell in grazie_ws[7]]
        assert grazie_ws.cell(excel_row, grazie_headers.index("Cacio") + 1).value == 1
        assert "Carzuc" not in grazie_headers
        assert "Amat" not in grazie_headers
        assert grazie_ws.cell(excel_row, grazie_headers.index("Altro") + 1).value == 2
        assert grazie_ws.cell(
            excel_row, grazie_headers.index("TOT PIATTI") + 1
        ).value == 3

        totals = wb["Totali"]
        totals_headers = [cell.value for cell in totals[1]]
        totals_row = 2 + 14
        assert totals.cell(totals_row, totals_headers.index("FLAMINIO") + 1).value == 4
        assert totals.cell(totals_row, totals_headers.index("GRAZIE") + 1).value == 3
        assert totals.cell(totals_row, totals_headers.index("TOTALI") + 1).value == 7
    finally:
        await client.drop_database(db_name)
