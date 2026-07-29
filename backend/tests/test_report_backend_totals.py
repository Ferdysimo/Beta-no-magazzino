import asyncio
import sys
from datetime import datetime
from pathlib import Path

import pytest
from fastapi import HTTPException
from openpyxl import Workbook


BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

import app.services.analysis as analysis_service
import app.services.report_snapshots as report_snapshots_service
from server import (
    _analysis_row_integrity,
    _compute_cash_sera_full,
    _compute_paste_breakdown_for_export,
    _compute_cash_sera_full_legacy_manual_prices,
    _compute_cassetto_total,
    _compute_paste_total_eur,
    _compute_paste_unrecognized,
    _collect_cursor_documents,
    _ensure_analysis_integrity,
    _get_daily_order_count,
    _pasta_dict_from_snapshot,
    _pasta_dict_snapshot_fields,
    _paste_text_from_order_docs,
    _prefetch_analysis_order_data,
    _should_create_pasta_dict_snapshot,
    _snapshot_report_paste_text_for_date,
    _write_analysis_locale_sheet,
    _write_totali_sheet_for_analysis,
)


PASTA_DICT = {"CARB": 8, "AMAT": 8}
PASTE_TEXT = "1 CARB\n2 X UNKNOWN\n3 AMAT"


class _FakeCursor:
    def __init__(self, documents):
        self._documents = list(documents)
        self._index = 0

    def __aiter__(self):
        self._index = 0
        return self

    async def __anext__(self):
        if self._index >= len(self._documents):
            raise StopAsyncIteration
        document = self._documents[self._index]
        self._index += 1
        return document


class _RangeCursor:
    def __init__(self, count):
        self._count = count
        self._index = 0

    def __aiter__(self):
        return self

    async def __anext__(self):
        if self._index >= self._count:
            raise StopAsyncIteration
        document = {"index": self._index}
        self._index += 1
        return document


class _FakeCollection:
    def __init__(self, documents=None, count=0):
        self.documents = list(documents or [])
        self.count = count
        self.find_queries = []
        self.count_queries = []

    def find(self, query, projection):
        self.find_queries.append(query)
        timestamp_field = next(key for key in query if key != "restaurant_id")
        bounds = query[timestamp_field]
        allowed_ids = set(query["restaurant_id"].get("$in", []))
        matches = [
            document
            for document in self.documents
            if document.get("restaurant_id") in allowed_ids
            and bounds["$gte"] <= document.get(timestamp_field, "") < bounds["$lt"]
        ]
        return _FakeCursor(matches)

    async def count_documents(self, query):
        self.count_queries.append(query)
        return self.count


class _FakeDatabase:
    def __init__(self, collections):
        self.collections = collections

    def __getitem__(self, name):
        return self.collections[name]


class _SnapshotCashCollection:
    def __init__(self, document):
        self.document = document
        self.updates = []

    async def find_one(self, query, projection):
        return dict(self.document)

    async def update_one(self, query, update, upsert=False):
        self.updates.append({"query": query, "update": update, "upsert": upsert})


def test_paste_total_uses_normalized_line_key_for_manual_prices():
    manual_prices = {"2 X UNKNOWN": "5"}

    assert _compute_paste_total_eur(PASTE_TEXT, manual_prices, PASTA_DICT) == 21


def test_paste_total_keeps_legacy_index_fallback_for_manual_prices():
    manual_prices = {"1": "5"}

    assert _compute_paste_total_eur(PASTE_TEXT, manual_prices, PASTA_DICT) == 21


def test_unrecognized_paste_reports_manual_price_from_normalized_line_key():
    manual_prices = {"2 X UNKNOWN": "5"}

    rows = _compute_paste_unrecognized(PASTE_TEXT, manual_prices, PASTA_DICT)

    assert rows == [{"idx": 1, "text": "2 X UNKNOWN", "manual_price": 5}]


def test_export_paste_breakdown_counts_types_and_altro():
    manual_prices = {"2 X UNKNOWN": "5"}

    result = _compute_paste_breakdown_for_export(PASTE_TEXT, manual_prices, PASTA_DICT)

    assert result["breakdown"]["CARB"]["count"] == 1
    assert result["breakdown"]["AMAT"]["count"] == 1
    assert result["unrecognized_count"] == 1
    assert result["total_count"] == 3
    assert result["total_eur"] == 21


def test_export_separates_known_analytical_type_without_list_price_from_altro():
    paste_text = "1 TONNO\n2 X UNKNOWN\n3 TONNO TA"
    manual_prices = {
        "1 TONNO": "8",
        "2 X UNKNOWN": "5",
        "3 TONNO TA": "8",
    }

    result = _compute_paste_breakdown_for_export(
        paste_text,
        manual_prices,
        PASTA_DICT,
    )

    assert "TONNO" not in PASTA_DICT
    assert result["breakdown"]["TONNO"] == {
        "count": 2,
        "total": 16,
        "price": 8,
    }
    assert result["unrecognized_count"] == 1
    assert result["unrecognized_eur"] == 5
    assert result["total_count"] == 3
    assert result["total_eur"] == 21


def test_export_groups_profile_excluded_pastas_in_altro_without_losing_revenue():
    result = _compute_paste_breakdown_for_export(
        "1 CARB\n2 AMAT\n3 CARZUC\n4 TONNO",
        {"4 TONNO": "8"},
        {"CARB": 8, "AMAT": 8, "CARZUC": 8},
        analysis_service.ANALYSIS_STANDARD_PASTA_TYPES,
    )

    assert "AMAT" not in result["breakdown"]
    assert "CARZUC" not in result["breakdown"]
    assert result["breakdown"]["CARB"]["count"] == 1
    assert result["breakdown"]["TONNO"]["count"] == 1
    assert result["unrecognized_count"] == 2
    assert result["unrecognized_eur"] == 16
    assert result["total_count"] == 4
    assert result["total_eur"] == 32


def test_analysis_pasta_profile_matches_reference_workbook_by_location():
    assert analysis_service._analysis_pasta_types_for_restaurant(
        {"location": "Flaminio"}
    ) == analysis_service.ANALYSIS_EXTENDED_PASTA_TYPES
    assert analysis_service._analysis_pasta_types_for_restaurant(
        {"location": "Largo di Brazzà"}
    ) == analysis_service.ANALYSIS_STANDARD_PASTA_TYPES
    assert analysis_service._analysis_pasta_types_for_restaurant(
        {"location": "Nuovo locale", "analysis_pasta_types": ["CARB", "AMATRICIANA"]}
    ) == ["CARB", "AMAT"]


def test_export_canonicalizes_long_pasta_alias_and_keeps_its_configured_price():
    result = _compute_paste_breakdown_for_export(
        "1 TARTUFO",
        {},
        {"TARTUFO": 9},
    )

    assert result["breakdown"]["TART"] == {
        "count": 1,
        "total": 9,
        "price": 9,
    }
    assert "TARTUFO" not in result["breakdown"]
    assert result["unrecognized_count"] == 0
    assert result["total_eur"] == 9


@pytest.mark.parametrize(
    ("line", "expected_sigla"),
    [
        ("42  CARB", "CARB"),
        ("42  carb senza pepe", "CARB"),
        ("7 AMAT - asporto", "AMAT"),
        ("9 CARZUC", "CARZUC"),
        ("10 - CARB", None),
        ("11 PIETRO CARB", None),
        ("12 CARB XL", None),
        ("13 SCONOSCIUTA", None),
    ],
)
def test_export_pasta_type_recognition_matches_report_rules(line, expected_sigla):
    dictionary = {"CARB": 8, "AMAT": 8, "CARZUC": 8}

    result = _compute_paste_breakdown_for_export(line, {}, dictionary)

    assert result["total_count"] == 1
    assert result["unrecognized_count"] == (1 if expected_sigla is None else 0)
    for sigla, values in result["breakdown"].items():
        assert values["count"] == (1 if sigla == expected_sigla else 0)


def test_cash_sera_legacy_manual_price_total_can_be_reconciled():
    cash_row = {
        "mattina": "100",
        "paste_text": PASTE_TEXT,
        "manual_prices": {"2 X UNKNOWN": "395"},
    }

    assert _compute_cash_sera_full_legacy_manual_prices(cash_row, [], PASTA_DICT) == 116
    assert _compute_cash_sera_full(cash_row, [], PASTA_DICT) == 511


def test_paste_text_from_order_docs_matches_report_frontend_format():
    docs = [
        {"order_number": 3, "description": " AMAT "},
        {"order_number": 1, "description": "CARB"},
        {"order_number": 2, "description": ""},
    ]

    assert _paste_text_from_order_docs(docs) == "1  CARB\n3  AMAT"


def test_analysis_prefetch_reads_valid_order_sources_and_deduplicates(monkeypatch):
    collections = {
        "orders": _FakeCollection([{
            "id": "order-1",
            "restaurant_id": "r1",
            "created_at": "2026-01-02T10:00:00+00:00",
            "order_number": 1,
            "description": "CARB",
        }]),
        "archived_orders": _FakeCollection([
            {
                "id": "order-1-copy",
                "restaurant_id": "r1",
                "created_at": "2026-01-02T10:00:00+00:00",
                "order_number": 1,
                "description": "CARB DUPLICATA",
            },
            {
                "id": "order-2",
                "restaurant_id": "r1",
                "created_at": "2026-01-02T11:00:00+00:00",
                "order_number": 2,
                "description": "AMAT",
            },
        ]),
        "deletion_logs": _FakeCollection([{
            "id": "delete-3",
            "restaurant_id": "r1",
            "original_created_at": "2026-01-02T12:00:00+00:00",
            "order_number": 3,
            "description": "PESTO",
        }]),
        "archived_deletion_logs": _FakeCollection([{
            "id": "delete-4",
            "restaurant_id": "r1",
            "original_created_at": "2026-01-02T13:00:00+00:00",
            "order_number": 4,
            "description": "RAGU",
        }]),
    }
    monkeypatch.setattr(analysis_service, "db", _FakeDatabase(collections))

    result = asyncio.run(_prefetch_analysis_order_data(
        ["r1"],
        "2026-01-01T23:00:00+00:00",
        "2026-01-02T23:00:00+00:00",
    ))

    key = ("r1", "2026-01-02")
    assert result["counts"][key] == 2
    assert result["texts"][key].splitlines() == [
        "1  CARB",
        "2  AMAT",
    ]
    assert collections["orders"].find_queries
    assert collections["archived_orders"].find_queries
    assert not collections["deletion_logs"].find_queries
    assert not collections["archived_deletion_logs"].find_queries


def test_analysis_prefetch_keeps_reused_number_with_distinct_creation_time(monkeypatch):
    collections = {
        "orders": _FakeCollection([{
            "id": "order-new",
            "restaurant_id": "r1",
            "created_at": "2026-01-02T14:00:00+00:00",
            "order_number": 7,
            "description": "AMAT",
        }]),
        "archived_orders": _FakeCollection([{
            "id": "order-old",
            "restaurant_id": "r1",
            "created_at": "2026-01-02T10:00:00+00:00",
            "order_number": 7,
            "description": "CARB",
        }]),
        "deletion_logs": _FakeCollection(),
        "archived_deletion_logs": _FakeCollection(),
    }
    monkeypatch.setattr(analysis_service, "db", _FakeDatabase(collections))

    result = asyncio.run(_prefetch_analysis_order_data(
        ["r1"],
        "2026-01-01T23:00:00+00:00",
        "2026-01-02T23:00:00+00:00",
    ))

    key = ("r1", "2026-01-02")
    assert result["counts"][key] == 2
    assert result["texts"][key].splitlines() == ["7  AMAT", "7  CARB"]


def test_cursor_collection_has_no_100000_document_truncation():
    result = asyncio.run(_collect_cursor_documents(_RangeCursor(100_001)))

    assert len(result) == 100_001
    assert result[-1]["index"] == 100_000


def test_daily_order_count_excludes_cancellations_and_keeps_reused_valid_numbers(monkeypatch):
    collections = {
        "orders": _FakeCollection([{
            "id": "active-1",
            "restaurant_id": "r1",
            "created_at": "2026-01-02T10:00:00+00:00",
            "order_number": 1,
            "description": "CARB",
        }]),
        "archived_orders": _FakeCollection([
            {
                "id": "active-1-copy",
                "restaurant_id": "r1",
                "created_at": "2026-01-02T10:00:00Z",
                "order_number": 1,
                "description": "CARB",
            },
            {
                "id": "reused-1",
                "restaurant_id": "r1",
                "created_at": "2026-01-02T11:00:00+00:00",
                "order_number": 1,
                "description": "AMAT",
            },
        ]),
        "deletion_logs": _FakeCollection([{
            "id": "delete-2",
            "restaurant_id": "r1",
            "original_created_at": "2026-01-02T12:00:00+00:00",
            "order_number": 2,
            "description": "PESTO",
        }]),
        "archived_deletion_logs": _FakeCollection([{
            "id": "delete-2-copy",
            "restaurant_id": "r1",
            "original_created_at": "2026-01-02T12:00:00Z",
            "order_number": 2,
            "description": "PESTO",
        }]),
    }
    monkeypatch.setattr(analysis_service, "db", _FakeDatabase(collections))

    result = asyncio.run(_get_daily_order_count(
        "r1",
        "2026-01-02T00:00:00+00:00",
        "2026-01-03T00:00:00+00:00",
    ))

    assert result == 2
    for name in ("orders", "archived_orders"):
        collection = collections[name]
        query = collection.find_queries[0]
        assert query["created_at"] == {
            "$gte": "2026-01-02T00:00:00+00:00",
            "$lt": "2026-01-03T00:00:00+00:00",
        }
    assert not collections["deletion_logs"].find_queries
    assert not collections["archived_deletion_logs"].find_queries


def test_pasta_dictionary_snapshot_is_versioned_and_not_recreated_for_history():
    fields = _pasta_dict_snapshot_fields(
        {"CARB": 8, "AMAT": 9},
        source="live_report",
        captured_at="2026-07-13T12:00:00+00:00",
    )

    assert _pasta_dict_from_snapshot(fields["pasta_dict_snapshot"]) == {"CARB": 8, "AMAT": 9}
    assert fields["pasta_dict_snapshot_version"] == 1
    assert fields["pasta_dict_snapshot_source"] == "live_report"
    assert _should_create_pasta_dict_snapshot(historical=False, existing_snapshot=None) is True
    assert _should_create_pasta_dict_snapshot(
        historical=False,
        existing_snapshot=fields["pasta_dict_snapshot"],
    ) is False
    assert _should_create_pasta_dict_snapshot(historical=True, existing_snapshot=None) is False


def test_midnight_snapshot_preserves_existing_dictionary(monkeypatch):
    existing_snapshot = _pasta_dict_snapshot_fields(
        {"CARB": 8},
        source="live_report",
        captured_at="2026-07-12T12:00:00+00:00",
    )
    cash = _SnapshotCashCollection({
        "paste_text": "1  CARB",
        "paste_manual_override": False,
        **existing_snapshot,
    })

    async def fake_build_paste_text(restaurant_id, date_rome_str):
        return "1  CARB\n2  AMAT"

    async def fail_if_dictionary_is_reloaded(restaurant_id):
        raise AssertionError("Lo snapshot esistente non deve essere sostituito")

    monkeypatch.setattr(
        report_snapshots_service,
        "db",
        type("Db", (), {"cash_daily_counts": cash})(),
    )
    monkeypatch.setattr(
        report_snapshots_service,
        "_build_paste_text_for_date",
        fake_build_paste_text,
    )
    monkeypatch.setattr(
        report_snapshots_service,
        "_get_pasta_dict_for",
        fail_if_dictionary_is_reloaded,
    )

    result = asyncio.run(_snapshot_report_paste_text_for_date(
        "2026-07-12",
        restaurant_ids=["r1"],
        snapshot_source="midnight",
    ))

    assert result["updated"] == 1
    set_fields = cash.updates[0]["update"]["$set"]
    assert set_fields["paste_text"] == "1  CARB\n2  AMAT"
    assert "pasta_dict_snapshot" not in set_fields
    assert "pasta_dict_snapshot_version" not in set_fields


def test_analysis_integrity_blocks_automatic_count_mismatch():
    integrity = _analysis_row_integrity(
        location="Flaminio",
        date_str="2026-01-02",
        source_count=4,
        paste_total_count=3,
        manual_override=False,
        has_snapshot=True,
        stored_paste_text="1 CARB\n2 AMAT\n3 RAGU",
        source_paste_text="1 CARB\n2 AMAT\n3 RAGU\n4 PESTO",
    )

    assert [item["code"] for item in integrity["errors"]] == ["paste_count_mismatch"]
    with pytest.raises(HTTPException) as exc:
        _ensure_analysis_integrity({"integrity": integrity})
    assert exc.value.status_code == 409
    assert exc.value.detail["issues"][0]["date"] == "2026-01-02"


def test_analysis_integrity_allows_manual_override_and_reports_missing_snapshot():
    integrity = _analysis_row_integrity(
        location="Grazie",
        date_str="2026-01-03",
        source_count=4,
        paste_total_count=3,
        manual_override=True,
        has_snapshot=False,
        stored_paste_text="1 CARB\n2 AMAT\n3 RAGU",
        source_paste_text="1 CARB\n2 AMAT\n3 RAGU\n4 PESTO",
    )

    assert integrity["errors"] == []
    assert {item["code"] for item in integrity["warnings"]} == {
        "manual_override_used",
        "manual_override_count_mismatch",
        "pasta_snapshot_missing",
    }
    _ensure_analysis_integrity({"integrity": integrity})


def test_analysis_integrity_reports_manual_override_even_when_count_matches():
    integrity = _analysis_row_integrity(
        location="Grazie",
        date_str="2026-01-04",
        source_count=2,
        paste_total_count=2,
        manual_override=True,
        has_snapshot=True,
        stored_paste_text="1 AMAT\n2 AMAT",
        source_paste_text="1 CARB\n2 CARB",
    )

    assert integrity["errors"] == []
    assert [item["code"] for item in integrity["warnings"]] == ["manual_override_used"]


def test_totali_sheet_reuses_deduplicated_source_counts():
    wb = Workbook()
    wb.remove(wb.active)
    restaurants = [{
        "id": "r1",
        "location": "Flaminio",
        "rows": [{
            "date_str": "2026-01-01",
            "source_order_count": 4,
        }],
    }]

    _write_totali_sheet_for_analysis(wb, restaurants, 2026)

    ws = wb["Totali"]
    assert ws["B2"].value == 4
    assert ws["C2"].value == 4


def test_cassetto_total_uses_loose_coin_values():
    row = {"cd5": "2", "cd2": "3", "cd1": "4", "cd05": "5"}

    assert _compute_cassetto_total(row) == 22.5


def test_analysis_locale_sheet_matches_reference_structure_and_style():
    wb = Workbook()
    wb.remove(wb.active)
    rest_data = {
        "location": "Flaminio",
        "pasta_dict": {"RAGU": 8, "AMAT": 8},
        "rows": [{
            "date": datetime(2026, 1, 1),
            "paste": {
                "breakdown": {
                    "RAGU": {"count": 2, "price": 8, "total": 16},
                    "AMAT": {"count": 1, "price": 8, "total": 8},
                },
                "unrecognized_count": 1,
                "unrecognized_eur": 5,
            },
            "paste_total_count": 4,
            "paste_total_eur": 29,
            "beverages": {
                sigla: {
                    "mattina": 10,
                    "inUsc": 2,
                    "scarti": 1,
                    "sera": 4,
                    "qty": 7,
                    "price": 2.5,
                    "incasso": 17.5,
                }
                for sigla in ("AL", "AG")
            },
            "bev_total_inc": 35,
            "cash": {"mattina": 100, "altro": 10},
            "spicci_total": 50,
            "cassetto_total": 22.5,
            "cash_sera": 146,
        }],
    }
    data = {"bev_sigle": ["AL", "AG"]}

    _write_analysis_locale_sheet(wb, rest_data, data, set())
    ws = wb["Flaminio"]
    row6 = [cell.value for cell in ws[6]]
    row7 = [cell.value for cell in ws[7]]

    assert ws["B2"].value.startswith("IN QUESTO FOGLIO VANNO MODIFICATE A MANO")
    assert ws["B4"].value == "PIATTI"
    assert "BEVANDE" in [cell.value for cell in ws[4]]
    assert "OUTPUT AUTOMATICO (n. Piatti)" in row6
    assert "Prezzi e incassi" in row6
    assert "SCARICHI" in row6
    assert "Altri utilizzi / scarti" in row6
    assert "INCASSI TOTALI" in row6
    assert "PASTE - NUMERO" not in row6
    assert "PASTE - INCASSO" not in row6
    assert "TOT EURO" not in row7
    assert "A L" in row7
    assert "A G" in row7
    assert "Cash in cassa mattina" in row7
    assert "Cash in cassa sera" in row7
    assert "€ 0,5" in row7

    price_group_start = row6.index("Prezzi e incassi") + 1
    assert ws.cell(8, price_group_start).value == 8
    assert ws.cell(8, price_group_start).fill.fgColor.rgb == "FFF4F5C1"
    assert ws.cell(8, price_group_start).number_format == "0"

    spicci_open_col = row7.index("Spicci aperti / portati") + 1
    cash_sera_col = row7.index("Cash in cassa sera") + 1
    assert ws.cell(8, spicci_open_col).value == 22.5
    assert ws.cell(8, spicci_open_col).number_format == "0.##"
    assert ws.cell(8, cash_sera_col).fill.fgColor.rgb == "FFF4F5C1"
    assert ws["B7"].font.name == "Calibri"
    assert ws["B8"].font.color.rgb == "FF000000"
    assert ws.sheet_view.showGridLines is True
    assert ws.sheet_view.zoomScale == 70


def test_analysis_locale_sheet_keeps_tonno_separate_from_altro():
    paste = _compute_paste_breakdown_for_export(
        "1 TONNO\n2 X UNKNOWN\n3 TONNO TA",
        {
            "1 TONNO": "8",
            "2 X UNKNOWN": "5",
            "3 TONNO TA": "8",
        },
        PASTA_DICT,
    )
    rest_data = {
        "location": "Flaminio",
        "pasta_dict": analysis_service._analysis_pasta_dict(PASTA_DICT),
        "rows": [{
            "date": datetime(2026, 7, 20),
            "paste": paste,
            "paste_total_count": paste["total_count"],
            "paste_total_eur": paste["total_eur"],
            "beverages": {},
            "bev_total_inc": 0,
            "cash": {},
            "spicci_total": 0,
            "cassetto_total": 0,
            "cash_sera": 0,
        }],
    }
    wb = Workbook()
    wb.remove(wb.active)

    _write_analysis_locale_sheet(
        wb,
        rest_data,
        {"bev_sigle": []},
        set(),
    )

    ws = wb["Flaminio"]
    headers = [cell.value for cell in ws[7]]
    assert headers[:11] == [
        "GIORNO", "Ragu", "Pesto", "Carb", "Cacio", "Pom",
        "Carzuc", "Tonno", "Tart", "Amat", "Altro",
    ]
    assert ws.cell(8, headers.index("Tonno") + 1).value == 2
    assert ws.cell(8, headers.index("Altro") + 1).value == 1
    assert ws.cell(8, headers.index("TOT PIATTI") + 1).value == 3
