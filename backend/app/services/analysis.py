import re
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Tuple

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.catalogs import BEVERAGES_CATALOG
from app.core.database import db
from app.core.time import (
    ROME_TZ,
    _rome_date_bounds_utc,
    _rome_date_from_iso,
)
from app.services.report import (
    ALL_CASH_FIELDS,
    PASTA_PRICES_MAP,
    _compute_cash_sera_full,
    _compute_cassetto_total,
    _compute_paste_total_eur,
    _compute_spicci_total,
    _eval_cash_value,
    _get_pasta_dict_for,
    _manual_price_for_paste_line,
    _pasta_dict_from_snapshot,
    _pasta_recognized_sigla,
)


def _paste_text_from_order_docs(order_docs: list) -> str:
    def sort_key(doc: dict):
        try:
            return int(doc.get("order_number") or 0)
        except Exception:
            return 0

    rows = []
    for doc in sorted(order_docs or [], key=sort_key):
        desc = (doc.get("description") or "").strip()
        if not desc:
            continue
        rows.append(f"{doc.get('order_number')}  {desc}")
    return "\n".join(rows)


ANALYSIS_ORDER_SOURCES: Tuple[Tuple[str, str], ...] = (
    ("orders", "created_at"),
    ("archived_orders", "created_at"),
)


ANALYSIS_DELETION_SOURCES: Tuple[Tuple[str, str], ...] = (
    ("deletion_logs", "original_created_at"),
    ("archived_deletion_logs", "original_created_at"),
)


def _canonical_analysis_order_timestamp(value) -> str:
    """Normalize equivalent ISO timestamps before using them as order identity."""
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc).isoformat(timespec="microseconds")
    except Exception:
        return str(value or "").strip()


def _normalize_analysis_order_doc(doc: dict, timestamp_field: str) -> Optional[dict]:
    restaurant_id = doc.get("restaurant_id")
    created_at = doc.get(timestamp_field)
    date_rome = _rome_date_from_iso(created_at)
    if not restaurant_id or not created_at or not date_rome:
        return None
    return {
        "id": doc.get("id"),
        "restaurant_id": restaurant_id,
        "created_at": created_at,
        "created_at_identity": _canonical_analysis_order_timestamp(created_at),
        "date_rome": date_rome,
        "order_number": doc.get("order_number"),
        "description": doc.get("description") or "",
    }


def _analysis_order_identity(doc: dict) -> tuple:
    # An order keeps its original creation timestamp when it is archived or
    # deleted. That timestamp distinguishes genuinely reused manual numbers,
    # while still collapsing copies left in two collections by an interrupted
    # archive operation.
    created_at_identity = doc.get("created_at_identity") or _canonical_analysis_order_timestamp(
        doc.get("created_at")
    )
    order_number = doc.get("order_number")
    if order_number not in (None, ""):
        return (
            doc.get("restaurant_id"),
            doc.get("date_rome"),
            created_at_identity,
            str(order_number),
        )
    return (
        doc.get("restaurant_id"),
        doc.get("date_rome"),
        created_at_identity,
        doc.get("id") or (doc.get("description") or "").strip(),
    )


async def _build_paste_text_for_date(
    restaurant_id: str,
    date_rome_str: str,
) -> str:
    start_utc, end_utc = _rome_date_bounds_utc(date_rome_str)
    source_data = await _prefetch_analysis_order_data(
        [restaurant_id],
        start_utc,
        end_utc,
    )
    return source_data["texts"].get((restaurant_id, date_rome_str), "")


async def _get_daily_order_count(
    restaurant_id: str,
    day_start: str,
    day_end_exclusive: str,
) -> int:
    source_data = await _prefetch_analysis_order_data(
        [restaurant_id],
        day_start,
        day_end_exclusive,
    )
    return sum(
        int(count or 0)
        for (rid, _date_rome), count in source_data["counts"].items()
        if rid == restaurant_id
    )


def _format_italian_long_date(day: datetime) -> str:
    weekdays = [
        "lunedi", "martedi", "mercoledi", "giovedi", "venerdi", "sabato", "domenica"
    ]
    months = [
        "gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
        "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"
    ]
    return f"{weekdays[day.weekday()]} {day.day} {months[day.month - 1]} {day.year}"


def _display_media_location(location: str) -> str:
    lowered = (location or "").lower()
    if "brazz" in lowered:
        return "BRAZZA"
    return (location or "").upper()


PREFERRED_PASTA_EXPORT_ORDER = [
    "RAGU", "PESTO", "CARB", "CACIO", "POM", "CARZUC",
    "TONNO", "TART", "TARTUFO", "AMAT", "AMATRICIANA",
]


ANALYSIS_STANDARD_PASTA_TYPES = [
    "RAGU", "PESTO", "CARB", "CACIO", "POM", "TART", "TONNO",
]


ANALYSIS_EXTENDED_PASTA_TYPES = [
    "RAGU", "PESTO", "CARB", "CACIO", "POM", "CARZUC",
    "TONNO", "TART", "AMAT",
]


ANALYSIS_PASTA_TYPES = ANALYSIS_EXTENDED_PASTA_TYPES


ANALYSIS_PASTA_ALIASES = {
    "TARTUFO": "TART",
    "AMATRICIANA": "AMAT",
}


PASTA_EXPORT_LABELS = {
    "RAGU": "Ragu",
    "PESTO": "Pesto",
    "CARB": "Carb",
    "CACIO": "Cacio",
    "POM": "Pom",
    "CARZUC": "Carzuc",
    "TONNO": "Tonno",
    "TART": "Tart",
    "TARTUFO": "Tart",
    "AMAT": "Amat",
    "AMATRICIANA": "Amat",
}


ANALYSIS_CASH_EXPORT_COLUMNS = [
    ("paste_total_eur", "PIATTI"),
    ("bev_total_inc", "BEVANDE"),
    ("altro", "ALTRO"),
    ("mattina", "Cash in cassa mattina"),
    ("sales_total", "TOTALE"),
    ("arr", "Arrotond."),
    ("vers", "Versam."),
    ("glo", "Glovo"),
    ("just", "Just Eat"),
    ("delv", "Deliveroo"),
    ("bp", "Buoni Pasto"),
    ("pos", "POS"),
    ("sat", "BP elett"),
    ("ft", "FT"),
    ("sp05", "€ 0,5"),
    ("sp1", "€ 1"),
    ("sp2", "€ 2"),
    ("sp5", "€ 5"),
    ("spicci_total", "Valori tubetti"),
    ("spicci_open", "Spicci aperti / portati"),
    ("cash_sera", "Cash in cassa sera"),
]


ANALYSIS_CASH_HEADER_FONT_SIZES = {
    "arr": 9,
    "vers": 9,
    "glo": 9,
    "just": 9,
    "sp05": 9,
    "sp1": 9,
    "sp2": 9,
    "sp5": 9,
    "spicci_total": 9,
    "spicci_open": 9,
    "cash_sera": 9,
}


def _validate_export_year(year: Optional[int]) -> int:
    selected_year = year or datetime.now(ROME_TZ).year
    if selected_year < 2020 or selected_year > 2100:
        raise HTTPException(status_code=400, detail="Anno non valido")
    return selected_year


def _analysis_year_days(selected_year: int) -> List[datetime]:
    current = datetime(selected_year, 1, 1, tzinfo=ROME_TZ)
    end = datetime(selected_year, 12, 31, tzinfo=ROME_TZ)
    days = []
    while current <= end:
        days.append(current)
        current += timedelta(days=1)
    return days


def _media_code_for_restaurant(restaurant: Dict) -> str:
    explicit_code = str(restaurant.get("report_code") or "").strip().upper()
    if explicit_code:
        return explicit_code
    display_name = _display_media_location(restaurant.get("location") or "")
    return display_name[:1].upper() if display_name else "?"


def _pasta_export_label(sigla: str) -> str:
    sigla_up = str(sigla or "").upper().strip()
    return PASTA_EXPORT_LABELS.get(sigla_up, sigla_up.title() if sigla_up else "")


def _ordered_pasta_dict(dict_map: Dict[str, float]) -> Dict[str, float]:
    clean = {str(k).upper().strip(): float(v or 0) for k, v in (dict_map or {}).items() if str(k).strip()}
    ordered: Dict[str, float] = {}
    for sigla in PREFERRED_PASTA_EXPORT_ORDER:
        if sigla in clean and sigla not in ordered:
            ordered[sigla] = clean[sigla]
    for sigla, price in clean.items():
        if sigla not in ordered:
            ordered[sigla] = price
    return ordered


def _analysis_pasta_types_for_restaurant(restaurant: Dict) -> List[str]:
    configured = restaurant.get("analysis_pasta_types")
    if isinstance(configured, (list, tuple)):
        normalized = []
        for value in configured:
            sigla = ANALYSIS_PASTA_ALIASES.get(
                str(value or "").upper().strip(),
                str(value or "").upper().strip(),
            )
            if sigla and sigla not in normalized:
                normalized.append(sigla)
        if normalized:
            return normalized

    location = str(restaurant.get("location") or "").strip().casefold()
    if location == "flaminio":
        return list(ANALYSIS_EXTENDED_PASTA_TYPES)
    return list(ANALYSIS_STANDARD_PASTA_TYPES)


def _analysis_pasta_dict(
    dict_map: Dict[str, float],
    pasta_types: Optional[List[str]] = None,
) -> Dict[str, Optional[float]]:
    """Return the analytical pasta categories independently from their prices."""
    priced = _ordered_pasta_dict(dict_map)
    analytical: Dict[str, Optional[float]] = {}
    selected_types = pasta_types or ANALYSIS_PASTA_TYPES
    selected_types = [
        ANALYSIS_PASTA_ALIASES.get(str(sigla).upper().strip(), str(sigla).upper().strip())
        for sigla in selected_types
        if str(sigla).strip()
    ]

    for sigla in selected_types:
        price = priced.get(sigla)
        if price is None:
            alias = next(
                (name for name, canonical in ANALYSIS_PASTA_ALIASES.items() if canonical == sigla and name in priced),
                None,
            )
            price = priced.get(alias) if alias else None
        analytical[sigla] = price

    for sigla, price in priced.items():
        canonical = ANALYSIS_PASTA_ALIASES.get(sigla, sigla)
        if canonical in ANALYSIS_PASTA_TYPES and canonical not in selected_types:
            continue
        if canonical in analytical:
            if analytical[canonical] is None:
                analytical[canonical] = price
            continue
        analytical[canonical] = price

    return analytical


def _compute_paste_breakdown_for_export(
    paste_text: str,
    manual_prices: Optional[dict] = None,
    dict_map: Optional[Dict[str, float]] = None,
    pasta_types: Optional[List[str]] = None,
) -> Dict:
    if dict_map is None:
        dict_map = PASTA_PRICES_MAP
    priced_dict = _ordered_pasta_dict(dict_map)
    analytical_dict = _analysis_pasta_dict(priced_dict, pasta_types)
    recognition_dict = dict(analytical_dict)
    for alias, canonical in ANALYSIS_PASTA_ALIASES.items():
        if canonical in analytical_dict:
            recognition_dict[alias] = analytical_dict.get(canonical)
    pricing_dict = dict(priced_dict)
    for alias, canonical in ANALYSIS_PASTA_ALIASES.items():
        if canonical in pricing_dict and alias not in pricing_dict:
            pricing_dict[alias] = pricing_dict[canonical]
    breakdown = {
        sigla: {"count": 0, "total": 0.0, "price": price}
        for sigla, price in analytical_dict.items()
    }
    inferred_prices: Dict[str, List[float]] = {sigla: [] for sigla in analytical_dict}
    unrecognized_count = 0
    unrecognized_eur = 0.0
    lines = [l.strip() for l in (paste_text or "").split("\n") if l.strip()]
    mp = manual_prices or {}

    for idx, line in enumerate(lines):
        recognized_sigla = _pasta_recognized_sigla(line, recognition_dict)
        priced_sigla = _pasta_recognized_sigla(line, pricing_dict)
        raw_manual_price = _manual_price_for_paste_line(mp, idx, line)
        try:
            manual_price = (
                float(str(raw_manual_price).replace(",", ".").strip())
                if str(raw_manual_price).strip()
                else 0.0
            )
        except Exception:
            manual_price = 0.0
        configured_line_price = pricing_dict.get(priced_sigla) if priced_sigla else None
        applied_price = (
            configured_line_price
            if configured_line_price is not None
            else max(manual_price, 0.0)
        )

        if recognized_sigla is not None:
            canonical_sigla = ANALYSIS_PASTA_ALIASES.get(recognized_sigla, recognized_sigla)
            configured_price = analytical_dict.get(canonical_sigla)
            breakdown.setdefault(
                canonical_sigla,
                {"count": 0, "total": 0.0, "price": configured_price},
            )
            breakdown[canonical_sigla]["count"] += 1
            breakdown[canonical_sigla]["total"] += applied_price
            if configured_price is None and manual_price > 0:
                inferred_prices.setdefault(canonical_sigla, []).append(manual_price)
            continue

        unrecognized_count += 1
        if applied_price > 0:
            unrecognized_eur += applied_price

    for sigla, prices in inferred_prices.items():
        unique_prices = {round(price, 6) for price in prices}
        if breakdown[sigla]["price"] is None and len(unique_prices) == 1:
            breakdown[sigla]["price"] = prices[0]
        breakdown[sigla]["total"] = round(breakdown[sigla]["total"], 2)

    recognized_count = sum(v["count"] for v in breakdown.values())
    recognized_eur = sum(v["total"] for v in breakdown.values())
    return {
        "breakdown": breakdown,
        "unrecognized_count": unrecognized_count,
        "unrecognized_eur": round(unrecognized_eur, 2),
        "total_count": recognized_count + unrecognized_count,
        "total_eur": round(recognized_eur + unrecognized_eur, 2),
    }


def _safe_sheet_title(title: str, used_titles: set) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", " ", str(title or "Locale")).strip() or "Locale"
    cleaned = re.sub(r"\s+", " ", cleaned)[:31]
    base = cleaned or "Locale"
    candidate = base
    idx = 2
    while candidate in used_titles:
        suffix = f" {idx}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        idx += 1
    used_titles.add(candidate)
    return candidate


async def _collect_cursor_documents(cursor) -> List[dict]:
    documents: List[dict] = []
    async for document in cursor:
        documents.append(document)
    return documents


async def _prefetch_analysis_source_data(
    restaurant_ids: List[str],
    start_utc: str,
    end_utc: str,
    sources: Tuple[Tuple[str, str], ...],
) -> Dict[str, Dict[tuple, object]]:
    grouped: Dict[tuple, Dict[tuple, dict]] = {}
    if not restaurant_ids:
        return {"texts": {}, "counts": {}}
    projection = {
        "_id": 0,
        "id": 1,
        "restaurant_id": 1,
        "created_at": 1,
        "original_created_at": 1,
        "order_number": 1,
        "description": 1,
    }
    for collection_name, timestamp_field in sources:
        query = {
            "restaurant_id": {"$in": restaurant_ids},
            timestamp_field: {"$gte": start_utc, "$lt": end_utc},
        }
        cursor = db[collection_name].find(query, projection)
        async for raw_doc in cursor:
            doc = _normalize_analysis_order_doc(raw_doc, timestamp_field)
            if not doc:
                continue
            group_key = (doc["restaurant_id"], doc["date_rome"])
            grouped.setdefault(group_key, {}).setdefault(_analysis_order_identity(doc), doc)
    texts: Dict[tuple, str] = {}
    counts: Dict[tuple, int] = {}
    for key, by_identity in grouped.items():
        docs = list(by_identity.values())
        texts[key] = _paste_text_from_order_docs(docs)
        counts[key] = len(docs)
    return {"texts": texts, "counts": counts}


async def _prefetch_analysis_order_data(
    restaurant_ids: List[str],
    start_utc: str,
    end_utc: str,
) -> Dict[str, Dict[tuple, object]]:
    """Load only orders that remain valid for production statistics."""
    return await _prefetch_analysis_source_data(
        restaurant_ids,
        start_utc,
        end_utc,
        ANALYSIS_ORDER_SOURCES,
    )


async def _prefetch_analysis_deleted_order_data(
    restaurant_ids: List[str],
    start_utc: str,
    end_utc: str,
) -> Dict[str, Dict[tuple, object]]:
    """Load cancellations only to disambiguate legacy automatic snapshots."""
    return await _prefetch_analysis_source_data(
        restaurant_ids,
        start_utc,
        end_utc,
        ANALYSIS_DELETION_SOURCES,
    )


def _normalized_paste_text(value: str) -> str:
    return "\n".join(line.strip() for line in (value or "").splitlines() if line.strip())


def _analysis_row_integrity(
    *,
    location: str,
    date_str: str,
    source_count: int,
    paste_total_count: int,
    manual_override: bool,
    has_snapshot: bool,
    stored_paste_text: str,
    source_paste_text: str,
) -> Dict[str, List[Dict]]:
    errors: List[Dict] = []
    warnings: List[Dict] = []

    def issue(code: str, message: str) -> Dict:
        return {
            "code": code,
            "location": location,
            "date": date_str,
            "expected_count": source_count,
            "actual_count": paste_total_count,
            "message": message,
        }

    if manual_override and (source_count > 0 or paste_total_count > 0):
        warnings.append(issue(
            "manual_override_used",
            "I tipi di pasta provengono da una correzione manuale salvata nel Report.",
        ))

    if source_count > 0 and paste_total_count != source_count:
        target = warnings if manual_override else errors
        target.append(issue(
            "manual_override_count_mismatch" if manual_override else "paste_count_mismatch",
            (
                "La forzatura manuale contiene un numero di paste diverso dagli ordini originali."
                if manual_override
                else "Il numero di paste ricostruite non coincide con gli ordini originali."
            ),
        ))
    elif source_count == 0 and paste_total_count > 0:
        warnings.append(issue(
            "source_orders_missing",
            "Sono presenti paste salvate nel Report ma non ordini sorgente negli archivi.",
        ))

    if paste_total_count > 0 and not has_snapshot:
        warnings.append(issue(
            "pasta_snapshot_missing",
            "Manca lo snapshot storico del dizionario paste; viene usato il dizionario attuale.",
        ))

    if (
        source_count > 0
        and not manual_override
        and _normalized_paste_text(stored_paste_text)
        and _normalized_paste_text(stored_paste_text) != _normalized_paste_text(source_paste_text)
    ):
        warnings.append(issue(
            "automatic_snapshot_rebuilt",
            "Lo snapshot automatico differiva dagli archivi ed è stato ricostruito per questo export.",
        ))

    return {"errors": errors, "warnings": warnings}


def _analysis_warning_counts(warnings: List[Dict]) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for warning in warnings:
        code = warning.get("code") or "unknown"
        counts[code] = counts.get(code, 0) + 1
    return counts


async def _build_annual_analysis_data(selected_year: int) -> Dict:
    restaurants = await _collect_cursor_documents(db.restaurants.find(
        {"role": "restaurant"},
        {"_id": 0, "password": 0},
    ))
    restaurants = sorted(restaurants, key=lambda r: (r.get("location") or r.get("username") or "").lower())
    restaurant_ids = [r.get("id") for r in restaurants if r.get("id")]
    days = _analysis_year_days(selected_year)
    start_date = f"{selected_year}-01-01"
    end_date = f"{selected_year}-12-31"
    start_utc, _ = _rome_date_bounds_utc(start_date)
    _, end_exclusive_utc = _rome_date_bounds_utc(end_date)

    cash_docs = await _collect_cursor_documents(db.cash_daily_counts.find(
        {"restaurant_id": {"$in": restaurant_ids}, "date_rome": {"$gte": start_date, "$lte": end_date}},
        {"_id": 0},
    ))
    cash_by_key = {(d.get("restaurant_id"), d.get("date_rome")): d for d in cash_docs}

    beverage_docs = await _collect_cursor_documents(db.beverage_daily_counts.find(
        {"restaurant_id": {"$in": restaurant_ids}, "date_rome": {"$gte": start_date, "$lte": end_date}},
        {"_id": 0},
    ))
    beverages_by_key: Dict[tuple, List[dict]] = {}
    for doc in beverage_docs:
        beverages_by_key.setdefault((doc.get("restaurant_id"), doc.get("date_rome")), []).append(doc)

    order_source_data = await _prefetch_analysis_order_data(
        restaurant_ids,
        start_utc,
        end_exclusive_utc,
    )
    source_paste_texts = order_source_data["texts"]
    source_order_counts = order_source_data["counts"]
    deletion_source_data = await _prefetch_analysis_deleted_order_data(
        restaurant_ids,
        start_utc,
        end_exclusive_utc,
    )
    deleted_order_counts = deletion_source_data["counts"]

    bev_sigle = [b["sigla"] for b in sorted(BEVERAGES_CATALOG, key=lambda x: x.get("sort_order", 999))]
    bev_prices = {b["sigla"]: b["price"] for b in BEVERAGES_CATALOG}
    restaurants_data = []
    integrity_errors: List[Dict] = []
    integrity_warnings: List[Dict] = []

    for rest in restaurants:
        rid = rest.get("id")
        price_dict = _ordered_pasta_dict(await _get_pasta_dict_for(rid))
        analysis_pasta_types = _analysis_pasta_types_for_restaurant(rest)
        analysis_dict = _analysis_pasta_dict(price_dict, analysis_pasta_types)
        summary_by_sigla = {sigla: 0 for sigla in analysis_dict.keys()}
        summary_unrecognized = 0
        summary_total = 0
        days_with_paste = 0
        rows = []

        for day in days:
            date_str = day.strftime("%Y-%m-%d")
            cash_doc = cash_by_key.get((rid, date_str), {}) or {}
            manual_prices = cash_doc.get("manual_prices") or {}
            stored_paste_text = cash_doc.get("paste_text") or ""
            source_paste_text = source_paste_texts.get((rid, date_str), "")
            source_order_count = int(source_order_counts.get((rid, date_str), 0) or 0)
            deleted_order_count = int(deleted_order_counts.get((rid, date_str), 0) or 0)
            manual_override = cash_doc.get("paste_manual_override") is True
            paste_text = (
                stored_paste_text
                if manual_override
                else (
                    source_paste_text
                    if source_order_count > 0 or deleted_order_count > 0
                    else stored_paste_text
                )
            )
            snapshot_dict = _pasta_dict_from_snapshot(cash_doc.get("pasta_dict_snapshot"))
            row_price_dict = _ordered_pasta_dict(
                snapshot_dict or price_dict
            )
            row_analysis_dict = _analysis_pasta_dict(
                row_price_dict,
                analysis_pasta_types,
            )
            for sigla, price in row_analysis_dict.items():
                if sigla not in analysis_dict or analysis_dict[sigla] is None:
                    analysis_dict[sigla] = price
                summary_by_sigla.setdefault(sigla, 0)
            paste_breakdown = _compute_paste_breakdown_for_export(
                paste_text,
                manual_prices,
                row_price_dict,
                analysis_pasta_types,
            )
            bev_docs = beverages_by_key.get((rid, date_str), [])
            bev_by_sigla = {d.get("sigla"): d for d in bev_docs}

            beverages = {}
            bev_total_qty = 0
            bev_total_inc = 0.0
            for sigla in bev_sigle:
                row = bev_by_sigla.get(sigla, {}) or {}
                m = _eval_cash_value(row.get("mattina"))
                u = _eval_cash_value(row.get("inUsc"))
                s = _eval_cash_value(row.get("scarti"))
                e = _eval_cash_value(row.get("sera"))
                qty = (0 if e == 0 else (m + u - e)) - s
                inc = round(qty * bev_prices.get(sigla, 0), 2)
                beverages[sigla] = {
                    "mattina": m,
                    "inUsc": u,
                    "scarti": s,
                    "sera": e,
                    "qty": int(qty),
                    "incasso": inc,
                    "price": bev_prices.get(sigla, 0),
                }
                bev_total_qty += int(qty)
                bev_total_inc += inc

            cash_calc = {**cash_doc, "paste_text": paste_text, "manual_prices": manual_prices}
            cash_values = {f: _eval_cash_value(cash_doc.get(f, "")) for f in ALL_CASH_FIELDS}
            spicci_total = round(_compute_spicci_total(cash_doc), 2)
            cassetto_total = round(_compute_cassetto_total(cash_doc), 2)
            cash_sera = round(_compute_cash_sera_full(cash_calc, bev_docs, row_price_dict), 2) if (cash_doc or paste_text or bev_docs) else 0.0

            total_count = paste_breakdown["total_count"]
            if total_count > 0:
                days_with_paste += 1
            summary_total += total_count
            summary_unrecognized += paste_breakdown["unrecognized_count"]
            for sigla, values in paste_breakdown["breakdown"].items():
                summary_by_sigla[sigla] = summary_by_sigla.get(sigla, 0) + values["count"]

            integrity = _analysis_row_integrity(
                location=rest.get("location") or rest.get("username") or "Locale",
                date_str=date_str,
                source_count=source_order_count,
                paste_total_count=total_count,
                manual_override=manual_override,
                has_snapshot=bool(snapshot_dict),
                stored_paste_text=stored_paste_text,
                source_paste_text=source_paste_text,
            )
            integrity_errors.extend(integrity["errors"])
            integrity_warnings.extend(integrity["warnings"])

            rows.append({
                "date": day,
                "date_str": date_str,
                "paste": paste_breakdown,
                "paste_total_count": total_count,
                "paste_total_eur": paste_breakdown["total_eur"],
                "beverages": beverages,
                "bev_total_qty": bev_total_qty,
                "bev_total_inc": round(bev_total_inc, 2),
                "cash": cash_values,
                "spicci_total": spicci_total,
                "cassetto_total": cassetto_total,
                "cash_sera": cash_sera,
                "has_report_data": bool(cash_doc or bev_docs or paste_text),
                "source_order_count": source_order_count,
                "deleted_order_count": deleted_order_count,
                "paste_manual_override": manual_override,
            })

        restaurants_data.append({
            "id": rid,
            "location": rest.get("location") or rest.get("username") or "Locale",
            "username": rest.get("username"),
            "report_code": rest.get("report_code") or "",
            "pasta_dict": analysis_dict,
            "rows": rows,
            "summary": {
                "total_paste": summary_total,
                "days_with_paste": days_with_paste,
                "unrecognized": summary_unrecognized,
                "by_sigla": summary_by_sigla,
            },
        })

    return {
        "year": selected_year,
        "days": days,
        "restaurants": restaurants_data,
        "bev_sigle": bev_sigle,
        "bev_prices": bev_prices,
        "integrity": {
            "errors": integrity_errors,
            "warnings": integrity_warnings,
            "warning_counts": _analysis_warning_counts(integrity_warnings),
        },
    }


def _apply_analysis_sheet_basics(ws):
    ws.freeze_panes = "B8"
    ws.sheet_view.showGridLines = True
    ws.sheet_view.zoomScale = 70


def _write_merged_group(
    ws,
    row_idx: int,
    start_col: int,
    end_col: int,
    label: str,
    fill_color: str,
    font_color: str = "000000",
):
    if start_col > end_col:
        return
    if start_col < end_col:
        ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
    cell = ws.cell(row_idx, start_col, label)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(
        name="Calibri",
        size=12,
        bold=label in ("VENDITE", "INCASSI TOTALI"),
        color=font_color,
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _analysis_group_fill(group: str) -> str:
    if not group:
        return "FFFFFFFF"
    if group == "OUTPUT AUTOMATICO (n. Piatti)":
        return "FFCCC1DA"
    if group == "Prezzi e incassi":
        return "FF604A7B"
    if group in (
        "MAGAZZ MATTINA",
        "SCARICHI",
        "Altri utilizzi / scarti",
        "MAGAZZ SERA",
        "VENDITE",
        "PREZZI",
        "INCASSO",
    ):
        return "FF77933C"
    if group == "INCASSI TOTALI":
        return "FF953735"
    return "FFE5E7EB"


def _analysis_group_font_color(group: str) -> str:
    if group == "OUTPUT AUTOMATICO (n. Piatti)":
        return "FF000000"
    if group == "Prezzi e incassi" or group in (
        "MAGAZZ MATTINA",
        "SCARICHI",
        "Altri utilizzi / scarti",
        "MAGAZZ SERA",
        "VENDITE",
        "PREZZI",
        "INCASSO",
        "INCASSI TOTALI",
    ):
        return "FFFFFFFF"
    return "FF000000"


def _analysis_header_fill(group: str) -> Optional[str]:
    return None


def _analysis_header_font_color(group: str) -> str:
    return "FF000000"


ANALYSIS_CASH_HEADER_STYLES = {
    "paste_total_eur": ("FF7030A0", "FFFFFFFF"),
    "bev_total_inc": ("FF4F6228", "FFFFFFFF"),
    "altro": ("FFFDEADA", "FF000000"),
    "mattina": ("FFD9D9D9", "FF000000"),
    "sales_total": ("FF984807", "FFFFFFFF"),
    "arr": ("FFFF0000", "FF000000"),
    "vers": ("FF3D85C6", "FFFFFFFF"),
    "glo": ("FFFFFFAF", "FF000000"),
    "just": ("FFFF9900", "FF000000"),
    "delv": ("FF00B050", "FFFFFFFF"),
    "bp": ("FFDCE6F2", "FF000000"),
    "sat": ("FFC4BD97", "FF000000"),
    "pos": ("FF7F7F7F", "FF000000"),
    "ft": ("FFF2F2F2", "FF000000"),
    "sp05": ("FFD0E0E3", "FF000000"),
    "sp1": ("FFFFF2CC", "FF000000"),
    "sp2": ("FFEAD1DC", "FF000000"),
    "sp5": ("FFD9EAD3", "FF000000"),
    "spicci_total": ("FFFCE5CD", "FF000000"),
    "spicci_open": ("FFE5B8B7", "FF000000"),
    "cash_sera": ("FFFFFF00", "FF000000"),
}


def _analysis_cash_header_style(field: str) -> Tuple[str, str]:
    return ANALYSIS_CASH_HEADER_STYLES.get(field, ("FF953735", "FFFFFFFF"))


def _analysis_body_fill(kind: str, group: str) -> Optional[str]:
    if kind == "paste_price":
        return "FFF4F5C1"
    if group in ("VENDITE", "PREZZI", "INCASSO"):
        return "FFF4F5C1"
    return None


def _analysis_cash_body_fill(field: str) -> Optional[str]:
    if field in ("sp05", "sp1", "sp2", "sp5"):
        return _analysis_cash_header_style(field)[0]
    if field == "cash_sera":
        return "FFF4F5C1"
    return None


def _analysis_short_pasta_label(sigla: str) -> str:
    sigla_up = str(sigla or "").upper().strip()
    if sigla_up in ("TART", "TARTUFO"):
        return "Tart"
    if sigla_up in ("AMAT", "AMATRICIANA"):
        return "Amat"
    if sigla_up == "CARZUC":
        return "Carzuc"
    return _pasta_export_label(sigla_up)


def _analysis_beverage_label(sigla: str) -> str:
    sigla_up = str(sigla or "").upper().strip()
    return {"AL": "A L", "AG": "A G"}.get(sigla_up, sigla_up)


def _analysis_money_number_format(value) -> str:
    try:
        numeric = float(value or 0)
    except Exception:
        return "0.##"
    if abs(numeric - round(numeric)) < 0.000001:
        return "0"
    return "0.##"


def _write_analysis_locale_sheet(wb: Workbook, rest_data: Dict, data: Dict, used_titles: set):
    title = _safe_sheet_title(rest_data["location"], used_titles)
    ws = wb.create_sheet(title=title)
    paste_siglas = list(rest_data["pasta_dict"].keys())
    bev_sigle = data["bev_sigle"]

    ws["B2"] = "IN QUESTO FOGLIO VANNO MODIFICATE A MANO SOLO LE CELLE DI QUESTO COLORE GIALLINO: Le altre son calcoli automatici"
    ws["B2"].font = Font(name="Calibri", size=12, bold=True, color="FF000000")

    columns = [{"group": "", "label": "GIORNO", "kind": "date"}]
    for sigla in paste_siglas:
        columns.append({"group": "OUTPUT AUTOMATICO (n. Piatti)", "label": _pasta_export_label(sigla), "kind": "paste_count", "sigla": sigla})
    columns.append({"group": "OUTPUT AUTOMATICO (n. Piatti)", "label": "Altro", "kind": "paste_unrecognized"})
    columns.append({"group": "OUTPUT AUTOMATICO (n. Piatti)", "label": "TOT PIATTI", "kind": "paste_total_count"})
    columns.append({"group": "", "label": "", "kind": "blank"})
    for sigla in paste_siglas:
        columns.append({"group": "Prezzi e incassi", "label": _analysis_short_pasta_label(sigla), "kind": "paste_price", "sigla": sigla})
    columns.append({"group": "Prezzi e incassi", "label": "", "kind": "blank"})
    for sigla in paste_siglas:
        columns.append({"group": "Prezzi e incassi", "label": _analysis_short_pasta_label(sigla), "kind": "paste_incasso", "sigla": sigla})
    columns.append({"group": "Prezzi e incassi", "label": "Altro", "kind": "paste_unrecognized_eur"})
    columns.append({"group": "", "label": "", "kind": "blank"})

    beverage_groups = [
        ("MAGAZZ MATTINA", "mattina"),
        ("SCARICHI", "inUsc"),
        ("Altri utilizzi / scarti", "scarti"),
        ("MAGAZZ SERA", "sera"),
        ("VENDITE", "qty"),
        ("PREZZI", "price"),
        ("INCASSO", "incasso"),
    ]
    for group_label, value_key in beverage_groups:
        for sigla in bev_sigle:
            columns.append({"group": group_label, "label": _analysis_beverage_label(sigla), "kind": "beverage", "sigla": sigla, "field": value_key})
        columns.append({"group": "", "label": "", "kind": "blank"})

    for key, label in ANALYSIS_CASH_EXPORT_COLUMNS:
        columns.append({"group": "INCASSI TOTALI", "label": label, "kind": "cash_export", "field": key})

    last_paste_col = max(
        [idx for idx, col in enumerate(columns, start=1) if col.get("group") in ("OUTPUT AUTOMATICO (n. Piatti)", "Prezzi e incassi")],
        default=2,
    )
    last_paste_price_col = max(
        [idx for idx, col in enumerate(columns, start=1) if col.get("kind") == "paste_price"],
        default=last_paste_col,
    )
    first_bev_col = next(
        (idx for idx, col in enumerate(columns, start=1) if col.get("group") == "MAGAZZ MATTINA"),
        None,
    )
    last_bev_col = max(
        [idx for idx, col in enumerate(columns, start=1) if col.get("kind") == "beverage"],
        default=first_bev_col or last_paste_col,
    )
    first_cash_col = next(
        (idx for idx, col in enumerate(columns, start=1) if col.get("kind") == "cash_export"),
        None,
    )

    for col_idx in range(2, len(columns) + 1):
        ws.cell(2, col_idx).fill = PatternFill("solid", fgColor="FFF4F5C1")
    if last_paste_col >= 2:
        for col_idx in range(2, last_paste_col + 1):
            ws.cell(4, col_idx).fill = PatternFill("solid", fgColor="FF7030A0")
        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=last_paste_price_col)
        ws.cell(4, 2, "PIATTI")
        ws.cell(4, 2).fill = PatternFill("solid", fgColor="FF7030A0")
        ws.cell(4, 2).font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
        ws.cell(4, 2).alignment = Alignment(horizontal="center", vertical="center")
    if first_bev_col:
        ws.merge_cells(start_row=4, start_column=first_bev_col, end_row=4, end_column=last_bev_col)
        ws.cell(4, first_bev_col, "BEVANDE")
        ws.cell(4, first_bev_col).fill = PatternFill("solid", fgColor="FF4F6228")
        ws.cell(4, first_bev_col).font = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")
        ws.cell(4, first_bev_col).alignment = Alignment(horizontal="center", vertical="center")

    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(7, idx, col["label"])
        group = col.get("group") or ""
        if col.get("kind") == "cash_export":
            fill_color, font_color = _analysis_cash_header_style(col.get("field") or "")
        else:
            fill_color, font_color = _analysis_header_fill(group), _analysis_header_font_color(group)
        if col.get("kind") == "paste_total_count":
            fill_color, font_color = "FF403152", "FFFFFFFF"
        if fill_color:
            cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(
            name="Calibri",
            size=ANALYSIS_CASH_HEADER_FONT_SIZES.get(col.get("field"), 12),
            bold=True,
            color=font_color,
        )
        if col.get("kind") == "paste_total_count":
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif col.get("kind") != "blank":
            cell.alignment = Alignment(vertical="bottom", wrap_text=True)

    group_start = None
    current_group = None
    for idx, col in enumerate(columns, start=1):
        group = col["group"]
        if group != current_group:
            if current_group:
                _write_merged_group(
                    ws,
                    6,
                    group_start,
                    idx - 1,
                    current_group,
                    _analysis_group_fill(current_group),
                    _analysis_group_font_color(current_group),
                )
            current_group = "" if group == "INCASSI TOTALI" else group
            group_start = idx
    if current_group:
        _write_merged_group(
            ws,
            6,
            group_start,
            len(columns),
            current_group,
            _analysis_group_fill(current_group),
            _analysis_group_font_color(current_group),
        )

    if first_cash_col:
        total_end_col = min(first_cash_col + 4, len(columns))
        _write_merged_group(
            ws,
            6,
            first_cash_col,
            total_end_col,
            "INCASSI TOTALI",
            "FF953735",
            "FFFFFFFF",
        )
        for col_idx in range(total_end_col + 1, len(columns) + 1):
            field = columns[col_idx - 1].get("field")
            fill = "FF632523" if field in ("pos", "sat", "ft", "sp05", "sp1", "sp2", "sp5", "spicci_total", "spicci_open", "cash_sera") else "FF953735"
            ws.cell(6, col_idx).fill = PatternFill("solid", fgColor=fill)

    money_kinds = {"paste_price", "paste_incasso", "paste_unrecognized_eur", "cash_export"}
    for row_idx, row_data in enumerate(rest_data["rows"], start=8):
        for col_idx, col in enumerate(columns, start=1):
            kind = col["kind"]
            value = None
            if kind == "date":
                value = row_data["date"].replace(tzinfo=None)
            elif kind == "paste_count":
                value = row_data["paste"]["breakdown"].get(col["sigla"], {}).get("count", 0)
            elif kind == "paste_unrecognized":
                value = row_data["paste"]["unrecognized_count"]
            elif kind == "paste_total_count":
                value = row_data["paste_total_count"]
            elif kind == "paste_price":
                value = row_data["paste"]["breakdown"].get(col["sigla"], {}).get("price", 0)
            elif kind == "paste_incasso":
                value = row_data["paste"]["breakdown"].get(col["sigla"], {}).get("total", 0)
            elif kind == "paste_unrecognized_eur":
                value = row_data["paste"]["unrecognized_eur"]
            elif kind == "paste_total_eur":
                value = row_data["paste_total_eur"]
            elif kind == "beverage":
                value = row_data["beverages"].get(col["sigla"], {}).get(col["field"], 0)
            elif kind == "cash_export":
                field = col["field"]
                if field == "paste_total_eur":
                    value = row_data["paste_total_eur"]
                elif field == "bev_total_inc":
                    value = row_data["bev_total_inc"]
                elif field == "sales_total":
                    value = round(
                        row_data["paste_total_eur"]
                        + row_data["bev_total_inc"]
                        + row_data["cash"].get("altro", 0),
                        2,
                    )
                elif field == "spicci_total":
                    value = row_data["spicci_total"]
                elif field == "spicci_open":
                    value = row_data.get("cassetto_total", 0)
                elif field == "cash_sera":
                    value = row_data["cash_sera"]
                else:
                    value = row_data["cash"].get(field, 0)
            cell = ws.cell(row_idx, col_idx, value if value not in (0, 0.0) else None)
            fill_color = _analysis_cash_body_fill(col.get("field") or "") if kind == "cash_export" else _analysis_body_fill(kind, col.get("group") or "")
            if fill_color:
                cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(
                name="Calibri",
                size=12,
                bold=kind == "cash_export" and col.get("field") == "sales_total",
                color="FF000000",
            )
            if kind == "date":
                cell.number_format = "dd/mm/yyyy"
            elif kind in money_kinds or (kind == "beverage" and col.get("field") in ("price", "incasso")):
                cell.number_format = _analysis_money_number_format(value)
            else:
                cell.number_format = "0"

    _apply_analysis_sheet_basics(ws)
    ws.sheet_format.defaultColWidth = 8.796875
    ws.column_dimensions["A"].width = 11.59765625
    cash_widths = {
        "paste_total_eur": 7,
        "bev_total_inc": 10.5,
        "altro": 8.5,
        "mattina": 12,
        "sales_total": 10.5,
        "pos": 7,
        "sat": 8,
        "ft": 7,
        "sp05": 5,
        "sp1": 5,
        "sp2": 5,
        "sp5": 5,
        "spicci_total": 8,
        "spicci_open": 10,
        "cash_sera": 10.5,
    }
    for col_idx, col in enumerate(columns[1:], start=2):
        width = 6.5
        if col.get("kind") == "cash_export":
            width = cash_widths.get(col.get("field"), 9)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[7].height = 46.8


def _write_totali_sheet_for_analysis(wb: Workbook, restaurants: List[Dict], selected_year: int):
    ws = wb.create_sheet(title="Totali")
    restaurants = sorted(restaurants, key=lambda r: (r.get("location") or "").lower())
    days = _analysis_year_days(selected_year)
    location_headers = [_display_media_location(r["location"]) for r in restaurants]
    media_headers = [f"MEDIA {_media_code_for_restaurant(r)}" for r in restaurants]
    headers = ["DATA", *location_headers, "TOTALI", *media_headers, "MEDIA T"]
    ws.append(headers)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    red_font = Font(color="D00000", name="Times New Roman", size=12)
    black_font = Font(color="000000", name="Times New Roman", size=12)
    header_font = Font(color="D00000", name="Times New Roman", size=12)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(color="000000", name="Times New Roman", size=12)

    counts_by_restaurant = {
        rest["id"]: {
            row["date_str"]: int(row.get("source_order_count", 0) or 0)
            for row in rest.get("rows", [])
        }
        for rest in restaurants
    }
    month_values = {r["location"]: [] for r in restaurants}
    today_rome = datetime.now(ROME_TZ)
    for day in days:
        date_str = day.strftime("%Y-%m-%d")
        excel_row = [_format_italian_long_date(day)]
        daily_total = 0
        for rest in restaurants:
            value = counts_by_restaurant.get(rest["id"], {}).get(date_str, 0)
            excel_row.append(value if value > 0 else None)
            daily_total += value
            if value > 0:
                month_values[rest["location"]].append(value)
        excel_row.append(daily_total if daily_total > 0 else None)

        is_month_end = (day + timedelta(days=1)).month != day.month
        is_completed_month = (
            day.year < today_rome.year
            or (day.year == today_rome.year and day.month < today_rome.month)
        )
        if is_month_end and is_completed_month:
            monthly_averages = []
            for rest in restaurants:
                values = month_values[rest["location"]]
                monthly_averages.append(round(sum(values) / len(values), 1) if values else None)
            valid_monthly_averages = [v for v in monthly_averages if v is not None]
            excel_row.extend(monthly_averages)
            excel_row.append(round(sum(valid_monthly_averages), 1) if valid_monthly_averages else None)
            month_values = {r["location"]: [] for r in restaurants}
        else:
            excel_row.extend([None] * (len(restaurants) + 1))
            if is_month_end:
                month_values = {r["location"]: [] for r in restaurants}

        ws.append(excel_row)

    max_col = len(headers)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=max_col):
        for idx, cell in enumerate(row, start=1):
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = red_font if 2 <= idx <= (1 + len(restaurants)) else black_font
            if idx >= len(restaurants) + 3 and cell.value is not None:
                cell.number_format = "0.0"

    ws.column_dimensions["A"].width = 34
    for col_idx in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    ws.freeze_panes = "A2"


def _analysis_summary_response(data: Dict) -> Dict:
    locations = []
    for rest in data["restaurants"]:
        by_sigla = rest["summary"]["by_sigla"]
        ordered = [
            {"sigla": sigla, "label": _pasta_export_label(sigla), "count": count}
            for sigla, count in by_sigla.items()
            if count > 0
        ]
        ordered.sort(key=lambda x: x["count"], reverse=True)
        locations.append({
            "id": rest["id"],
            "location": rest["location"],
            "total_paste": rest["summary"]["total_paste"],
            "days_with_paste": rest["summary"]["days_with_paste"],
            "unrecognized": rest["summary"]["unrecognized"],
            "top_paste": ordered[:8],
            "all_paste": ordered,
        })
    return {
        "year": data["year"],
        "locations": locations,
        "total_paste": sum(r["summary"]["total_paste"] for r in data["restaurants"]),
    }


def _ensure_analysis_integrity(data: Dict) -> None:
    errors = (data.get("integrity") or {}).get("errors") or []
    if not errors:
        return
    raise HTTPException(
        status_code=409,
        detail={
            "message": "Export bloccato: alcune giornate non coincidono con gli ordini sorgente.",
            "error_count": len(errors),
            "issues": errors[:20],
        },
    )
